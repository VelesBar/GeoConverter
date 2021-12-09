from openpyxl import load_workbook
import requests
import time
start_time = time.time()

##log_file = open('#log_file.txt', 'w')
##log_file.write('log start \n')

class GeoParcer(object):

    def API_get_responce(url,coordinate,iterator,method):
        if method == 0 or method == None:
            res = requests.get(url \
                               + coordinate[iterator] \
                               + '&kind=house' \
                               + '&results=1' \
                               + '&format=json')
            ##log_file.write('\n server response \n')
            ##log_file.write(str(iterator) + '\n')
            ##log_file.write(str(res))
            location = res.json()
            trigger_try = True
            return res
        else:
            res = requests.get(url \
                               + coordinate[iterator] \
                               + '&results=1' \
                               + '&format=json')
            #log_file.write('\n server response \n' )
            #log_file.write(str(iterator) + '\n')
            #log_file.write(str(res))
            trigger_try = True
            return res

    def read_excel(away):
        wb = load_workbook(away)
        book = wb['Лист1']
        column_a = book['A']
        column_b = book['B']
        coordinate_lat = []
        coordinate_lon = []
        coordinate = []
        for i in range(len(column_a)):
            coordinate_lat.append(column_a[i].value)
        for b in range(len(column_b)):
            coordinate_lon.append(column_b[b].value)
        if coordinate_lon[0] == 0 or None:
            b=0
            for b in range(len(coordinate_lat)):
                coordinate_lat[b], coordinate_lon[b] = coordinate_lat[b].split(', ')
                coordinate.append(str(coordinate_lon[b]) + ',' + str(coordinate_lat[b]))
        else:
            for c in range(len(coordinate_lat)):
                coordinate.append(str(coordinate_lon[c]) + ',' + str(coordinate_lat[c]))
                #log_file.write('\n Cordinate \n')
                #log_file.write(str(coordinate))
        return coordinate

    def get_address(res):
        location = res.json()
        #print(location)
        #log_file.write('\n location json \n')
        #log_file.write(str(location))

        try:
            location = location['response'] \
                ['GeoObjectCollection'] \
                ['featureMember'] \
                [0]['GeoObject'] \
                ['metaDataProperty'] \
                ['GeocoderMetaData'] \
                ['Address'] \
                ['Components']
            return location

        except:
            return 1

    def put_excel_file(away_out,row,location,get_method, menu_key):
        wb = load_workbook(away_out)
        book = wb['Sheet']
        sheet = wb.active
        def put_point(row, column, location):
            point_cell = sheet.cell(row=row + 1, column=column)
            point_cell.value = location

        point_c = 0
        point_p = 0
        count_p = 0
        point_a = 0
        point_l = 0
        point_s = 0
        point_h = 0
        point=0
        if get_method == 1:

            for i in range(len(location)):

                if location[i]['kind'] == 'country':
                    put_point(row, 4+point_c, location[i]['name'])
                    point_c+=1
                if location[i]['kind'] == 'province':
                    put_point(row, 6+point_p, location[i]['name'])
                    point_p+=1
                if location[i]['kind'] == 'area':
                    put_point(row, 8+point_a, location[i]['name'])
                    point_a+=1
                if location[i]['kind'] == 'locality':
                    put_point(row, 10+point_l, location[i]['name'])
                    point_l+=1
                if location[i]['kind'] == 'street':
                    put_point(row, 12+point_s, location[i]['name'])
                    point_s+=1
                if location[i]['kind'] == 'house':
                    put_point(row, 14+point_h, location[i]['name'])
                    point_h+=1
        if get_method == 2:
            '''
            for i in range(len(location)):
                for b in range(len(menu_key)):
                    if location[i]['kind'] == menu_key[b]:
                        put_point(row, i + 6 + point_a, location[i]['name'])
                        point_a += 1'''
            for i in range(len(location)):
                for b in range(len(menu_key)):
                    if (location[i]['kind'] == menu_key[b]) and (location[i]['kind'] == 'country'):
                            put_point(row, 4 + point_c, location[i]['name'])
                            point_c += 1

                    elif ((location[i]['kind'] == menu_key[b]) or (menu_key[b] == 'province_obl')) and (location[i]['kind'] == 'province'):
                            count_p +=1 #считаем вхождения по этому условию, чтобы не вставить лишнего
                            if (menu_key[b] == 'province_obl') and (count_p == 2):
                                put_point(row, 7 + point_p, location[i]['name'])
                            elif menu_key[b] == 'province':
                                put_point(row, 6 + point_p, location[i]['name'])

                    elif (location[i]['kind'] == menu_key[b]) and (location[i]['kind'] == 'area'):
                            put_point(row, 8 + point_a, location[i]['name'])
                            point_a += 1

                    elif (location[i]['kind'] == menu_key[b]) and (location[i]['kind'] == 'locality'):
                            put_point(row, 10 + point_l, location[i]['name'])
                            point_l += 1

                    elif (location[i]['kind'] == menu_key[b]) and (location[i]['kind'] == 'street'):
                            put_point(row, 12 + point_s, location[i]['name'])
                            point_s += 1

                    elif (location[i]['kind'] == menu_key[b]) and (location[i]['kind'] == 'house'):
                            put_point(row, 14 + point_h, location[i]['name'])
                            point_h += 1

        wb.save('out_file.xlsx')

        return 0



#Reading key
key_file = open('./key.txt')
key = key_file.read()
key = key.replace(' ', '')

#формируем url
url = 'https://geocode-maps.yandex.ru/1.x/?apikey='+key+'&geocode='
away = './Book.xlsx'
away_out = './out_file.xlsx'

'''
Нужный функционал: возможность выбирать, в какие именно данные преобразовывать координаты. 
Только улица, только город, только дом, или все вместе. 
A - вести всё
'''

#Меню выбора типа записи в файл
print('Выберите тип записи в файл:')
print('1. Вывести адрес полностью')
print('2. Настроить вывод вручную')
menu_first_input = input()
menu_key = []

if menu_first_input == '1':
    get_method = 1
else:
    get_method = 2
    print('Введите желаемые номера через запятую без пробелов. \nпример "1,2,3"\n' \
          '1: страна\n' \
          '2: федеральный округ\n' \
          '3: область\n'\
          '4: городско'
          'й округ\n' \
          '5: город\n' \
          '6: улица\n' \
          '7: номер дома\n' \
          )
    menu_arr = input().split(',')

    menu = {
    '1':'country', #страна
    '2':'province', #фед округ
    '3':'province_obl', #область
    '4':'area', #городской округ
    '5':'locality', #город
    '6':'street', #улица
    '7':'house' #номер дома
    }

    menu_key = []
    for i in range(len(menu_arr)):
        menu_key.append(menu[menu_arr[i]])
    print(menu_key)

#Основной код, вызов методов класса
coordinate = GeoParcer.read_excel(away)
for row in range(len(coordinate)):
    res = GeoParcer.API_get_responce(url,coordinate,row, 0)
    address = GeoParcer.get_address(res)
    if row%20==0:
        print('Обработано '+str(row)+' строк из '+str(len(coordinate)))
    #если метод get_address() возвращает 1, значит произошла ошибка. Необходимо передать в метод API_get_responce, method=1
    if address == 1:
        while address == 1:
            #print(res)
            #print(address)
            res = GeoParcer.API_get_responce(url, coordinate, row, 1)
            address = GeoParcer.get_address(res)
        for b in range(len(address)):
            GeoParcer.put_excel_file(away_out, row, address, get_method, menu_key)
    else:
        for b in range(len(address)):
            GeoParcer.put_excel_file(away_out, row, address, get_method, menu_key)
    #else:
    #for b in range(len(address)):
    #GeoParcer.put_excel_file(away_out, row, address, get_method, menu_key)

print("--- %s seconds ---" % (time.time() - start_time))
print("Для завершения работы программы, нажмите клавишу Enter")
#log_file.close()
input()

